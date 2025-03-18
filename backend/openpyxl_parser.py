import json
import os
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime
import yaml
import traceback

class ExcelSheetParser:
    """Parser for extracting structured data from Excel sheets based on a JSON template."""
    
    def __init__(self, excel_path, template_path=None):
        """
        Initialize the parser.
        
        Parameters:
        -----------
        excel_path : str
            Path to the Excel file to parse
        template_path : str, optional
            Path to a JSON template file that describes the structure
        """
        self.excel_path = excel_path
        self.template = None
        self.wb = None
        
        if template_path:
            self.load_template(template_path)
            
    def load_template(self, template_path):
        """Load a JSON template file."""
        print(f"Loading template from {template_path}")
        with open(template_path, 'r', encoding='utf-8') as f:
            self.template = json.load(f)
        return self.template
            
    def open_workbook(self):
        """Open the Excel workbook."""
        if not self.wb:
            print(f"Opening workbook: {self.excel_path}")
            self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        return self.wb
        
    def get_sheet_names(self):
        """Get all sheet names in the workbook."""
        wb = self.open_workbook()
        return wb.sheetnames
        
    def find_data_region(self, sheet_name):
        """
        Find the main data region in a sheet based on template info or auto-detection.
        
        Returns:
        --------
        tuple
            (start_row, title_col, data_start_col)
        """
        wb = self.open_workbook()
        ws = wb[sheet_name]
        
        # If we have template data, use that
        if self.template and 'data' in self.template:
            metadata = self.template.get('metadata', {})
            if metadata.get('sheet') == sheet_name:
                data_region = self.template.get('data', {}).get('assets', {}).get('data_region', {})
                if data_region:
                    return (
                        data_region.get('start_row', 10),
                        data_region.get('title_col', 1),
                        data_region.get('data_start_col', 2)
                    )
        
        # Otherwise, auto-detect
        # Look for rows with significant data (not empty)
        for row_idx in range(1, min(50, ws.max_row)):  # Check first 50 rows
            non_empty_count = 0
            for col_idx in range(1, min(20, ws.max_column)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None and str(cell_value).strip():
                    non_empty_count += 1
                    
            # If we found a row with at least 5 non-empty cells, consider it the data start
            if non_empty_count >= 5:
                # First column is typically for titles
                title_col = 1
                # Data typically starts from the 2nd column
                data_start_col = 2
                return (row_idx, title_col, data_start_col)
                
        # Default values if nothing was found
        return (10, 1, 2)
        
    def extract_dates_from_header(self, sheet_name, data_region=None):
        """
        Extract dates from the header row of the data region.
        
        Parameters:
        -----------
        sheet_name : str
            Name of the sheet to extract dates from
        data_region : tuple, optional
            (start_row, title_col, data_start_col)
            
        Returns:
        --------
        list
            List of datetime objects
        """
        if data_region is None:
            data_region = self.find_data_region(sheet_name)
            
        start_row, title_col, data_start_col = data_region
        wb = self.open_workbook()
        ws = wb[sheet_name]
        
        print(f"  Extracting dates from header row {start_row-1}, starting at column {data_start_col}")
        
        dates = []
        for col in range(data_start_col, ws.max_column + 1):
            cell_value = ws.cell(row=start_row-1, column=col).value
            
            # Print info about what we're seeing
            print(f"    Column {col}: {type(cell_value)} - {cell_value}")
            
            # Skip empty cells
            if cell_value is None:
                continue
                
            # Try to parse as date (could be a date object or string)
            date_value = None
            if isinstance(cell_value, datetime):
                date_value = cell_value
            else:
                # Try common date formats
                try:
                    date_value = pd.to_datetime(cell_value)
                except:
                    pass
                
            if date_value:
                dates.append(date_value)
                
        print(f"  Found {len(dates)} dates in header")
        
        # If no dates found but this is the sheet from the template, use those dates
        if not dates and self.template:
            metadata = self.template.get('metadata', {})
            if metadata.get('sheet') == sheet_name:
                template_dates = metadata.get('dates', [])
                if template_dates:
                    print(f"  Using {len(template_dates)} dates from template")
                    return [pd.to_datetime(d) for d in template_dates]
        
        # If still no dates, create some sample dates for structure
        if not dates:
            print("  No dates found, creating sample dates")
            # Create monthly dates for the last 2 years
            end_date = datetime.now()
            dates = [end_date - pd.DateOffset(months=i) for i in range(24)]
            dates.reverse()  # Oldest first
            
        return dates
    
    def detect_hierarchy(self, sheet_name, data_region=None):
        """
        Detect hierarchy from indentation in the title column.
        
        Parameters:
        -----------
        sheet_name : str
            Name of the sheet to analyze
        data_region : tuple, optional
            (start_row, title_col, data_start_col)
            
        Returns:
        --------
        dict
            Hierarchical structure of the data
        """
        if data_region is None:
            data_region = self.find_data_region(sheet_name)
            
        start_row, title_col, data_start_col = data_region
        wb = self.open_workbook()
        ws = wb[sheet_name]
        
        print(f"  Detected data region: start_row={start_row}, title_col={title_col}, data_start_col={data_start_col}")
        
        # Get dates from header
        dates = self.extract_dates_from_header(sheet_name, data_region)
        date_strs = [d.strftime("%Y-%m-%d") for d in dates]
        
        # Initialize root nodes for assets and liabilities
        assets = {
            "name": "EIGNIR / ASSETS",
            "children": []
        }
        
        liabilities = {
            "name": "SKULDIR / LIABILITIES",
            "children": []
        }
        
        # Stack to keep track of the current hierarchy
        hierarchy_stack = []
        current_section = None
        
        # Variable to determine if we're in the assets or liabilities section
        in_assets = True
        
        print(f"  Processing rows from {start_row} to {ws.max_row}")
        
        # Track how many items we find
        assets_count = 0
        liabilities_count = 0
        
        # Iterate through rows
        for row_idx in range(start_row, ws.max_row + 1):
            title_cell = ws.cell(row=row_idx, column=title_col)
            title_text = title_cell.value
            
            # Skip empty rows
            if title_text is None or not str(title_text).strip():
                continue
                
            # Extract indentation level (based on leading spaces in Excel formula)
            title_text = str(title_text).strip()
            
            # Check if this is a switch between assets and liabilities
            if "SKULDIR" in title_text or "LIABILITIES" in title_text and row_idx > start_row + 5:
                in_assets = False
                hierarchy_stack = []  # Reset the stack
                print(f"  Switched to liabilities section at row {row_idx}")
                
            # Determine level based on cell's indentation or text
            level = 0
            if hasattr(title_cell, 'alignment') and title_cell.alignment and title_cell.alignment.indent:
                level = title_cell.alignment.indent
            else:
                # Try to guess based on leading spaces
                leading_spaces = len(title_text) - len(title_text.lstrip())
                level = leading_spaces // 2  # Assuming 2 spaces per indentation level
            
            # Create node for this item
            node = {
                "name": title_text,
                "is": title_text.split(" / ")[0] if " / " in title_text else title_text,
                "en": title_text.split(" / ")[1] if " / " in title_text else "",
                "values": {}
            }
            
            # Add values for each date
            for col_idx, date_str in enumerate(date_strs, start=data_start_col):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None and isinstance(value, (int, float)):
                    node["values"][date_str] = value
            
            # Adjust the hierarchy stack based on level
            while len(hierarchy_stack) > level:
                hierarchy_stack.pop()
                
            if level == 0:
                # This is a top-level item
                if in_assets:
                    assets["children"].append(node)
                    assets_count += 1
                else:
                    liabilities["children"].append(node)
                    liabilities_count += 1
                hierarchy_stack = [node]
            else:
                # This is a child item
                if not hierarchy_stack:
                    # If stack is empty but we have a child item, add it to the appropriate top level
                    if in_assets:
                        assets["children"].append(node)
                        assets_count += 1
                    else:
                        liabilities["children"].append(node)
                        liabilities_count += 1
                    hierarchy_stack = [node]
                else:
                    # Add as child to the current parent
                    parent = hierarchy_stack[-1]
                    if "children" not in parent:
                        parent["children"] = []
                    parent["children"].append(node)
                    hierarchy_stack.append(node)
        
        print(f"  Found {assets_count} assets items and {liabilities_count} liabilities items")
        
        # Create the result structure
        result = {
            "metadata": {
                "file": os.path.basename(self.excel_path),
                "sheet": sheet_name,
                "dates": date_strs,
                "extracted_at": datetime.now().isoformat()
            },
            "data": {
                "assets": assets,
                "liabilities": liabilities
            }
        }
        
        return result
    
    def parse_sheet(self, sheet_name):
        """
        Parse a specific sheet and extract its hierarchical structure.
        
        Parameters:
        -----------
        sheet_name : str
            Name of the sheet to parse
            
        Returns:
        --------
        dict
            Parsed data structure
        """
        data_region = self.find_data_region(sheet_name)
        return self.detect_hierarchy(sheet_name, data_region)
    
    def save_as_json(self, data, output_path):
        """Save parsed data as a JSON file."""
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print(f"Saved JSON to {output_path}")
        
    def save_as_yaml(self, data, output_path):
        """Save parsed data as a YAML file."""
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            yaml.dump(data, f, sort_keys=False, allow_unicode=True)
        print(f"Saved YAML to {output_path}")

def main():
    try:
        # Path to the Excel file and template JSON
        excel_path = "backend/cache/INN_ReikningarBankakerfis_012025.xlsx"
        template_path = "backend/parsed_data/banking_system_accounts.json"
        
        print(f"Starting Excel parser for {excel_path}")
        
        # Check if files exist
        if not os.path.exists(excel_path):
            print(f"ERROR: Excel file not found: {excel_path}")
            return
        
        if not os.path.exists(template_path):
            print(f"WARNING: Template file not found: {template_path}")
            template_path = None
        
        # Create parser
        parser = ExcelSheetParser(excel_path, template_path)
        
        # Get sheet names
        sheet_names = parser.get_sheet_names()
        print(f"Found sheets: {sheet_names}")
        
        for sheet_name in sheet_names:
            print(f"\nProcessing sheet: {sheet_name}")
            try:
                # Parse the sheet
                parsed_data = parser.parse_sheet(sheet_name)
                
                # Save as JSON and YAML
                output_json = f"backend/parsed_data/wb_parse_obj_{sheet_name}.json"
                output_yaml = f"backend/parsed_data/wb_parse_obj_{sheet_name}.yaml"
                
                parser.save_as_json(parsed_data, output_json)
                parser.save_as_yaml(parsed_data, output_yaml)
                
                # Print summary
                dates = parsed_data["metadata"]["dates"]
                assets = parsed_data["data"]["assets"]["children"]
                liabilities = parsed_data["data"]["liabilities"]["children"]
                
                if dates:
                    print(f"  - Found {len(dates)} dates from {dates[0]} to {dates[-1]}")
                else:
                    print(f"  - No dates found in the sheet")
                    
                print(f"  - Found {len(assets)} top-level assets and {len(liabilities)} top-level liabilities")
                
            except Exception as e:
                print(f"Error processing sheet {sheet_name}: {str(e)}")
                traceback.print_exc()
        
        print("\nProcessing complete.")
    except Exception as e:
        print(f"ERROR: {str(e)}")
        traceback.print_exc()

if __name__ == "__main__":
    main() 