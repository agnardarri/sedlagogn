import json
import os
import openpyxl
import pandas as pd
from datetime import datetime

def load_schema():
    """Load the schema from test.json"""
    schema_path = "backend/config/schemas/test.json"
    with open(schema_path, 'r', encoding='utf-8') as f:
        schema = json.load(f)
    return schema

def extract_hierarchy_data(excel_path):
    """Extract hierarchical data from the Excel file based on the schema"""
    schema = load_schema()
    
    # Get configuration for the banking system accounts
    banking_config = schema["INN"]["data"]["reikningar_bankakerfis"]["config"]["sheets"][0]
    
    # Open the Excel file
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_name = banking_config["sheet"]
    ws = wb[sheet_name]
    
    # Extract configuration parameters
    date_row = banking_config["date_row"]
    data_start_column = banking_config["data_start_column"]
    data_name_column = banking_config["data_name_column"]
    assets_row = banking_config["assets_row"]
    liabilities_row = banking_config["liabilities_row"]
    assets_hierarchy = banking_config["assets_hierarchy"]
    liabilities_hierarchy = banking_config["liabilities_hierarchy"]
    
    # Extract dates from header row
    dates = []
    for col in range(data_start_column, ws.max_column + 1):
        date_value = ws.cell(row=date_row, column=col).value
        if date_value:
            if isinstance(date_value, datetime):
                dates.append(date_value.strftime("%Y-%m-%d"))
            else:
                try:
                    dt = pd.to_datetime(date_value)
                    dates.append(dt.strftime("%Y-%m-%d"))
                except:
                    # If it's not a date, just skip it
                    pass
    
    # Extract assets data with hierarchy
    assets_data = []
    row_index = assets_row
    for level in assets_hierarchy:
        # Get cell value
        cell_value = ws.cell(row=row_index, column=data_name_column).value
        
        if cell_value is None:
            # Skip empty rows
            row_index += 1
            continue
            
        # Get values for this row
        values = {}
        for date_idx, col in enumerate(range(data_start_column, data_start_column + len(dates))):
            value = ws.cell(row=row_index, column=col).value
            if value is not None and isinstance(value, (int, float)):
                values[dates[date_idx]] = value
        
        # Create entry with hierarchy level
        entry = {
            "name": str(cell_value).strip(),
            "hierarchy_level": level,
            "row": row_index,
            "values": values
        }
        
        # Split name into Icelandic and English if available
        if " / " in entry["name"]:
            parts = entry["name"].split(" / ", 1)
            entry["is"] = parts[0]
            entry["en"] = parts[1]
        else:
            entry["is"] = entry["name"]
            entry["en"] = ""
            
        assets_data.append(entry)
        row_index += 1
    
    # Extract liabilities data with hierarchy
    liabilities_data = []
    row_index = liabilities_row
    for level in liabilities_hierarchy:
        # Get cell value
        cell_value = ws.cell(row=row_index, column=data_name_column).value
        
        if cell_value is None:
            # Skip empty rows
            row_index += 1
            continue
            
        # Get values for this row
        values = {}
        for date_idx, col in enumerate(range(data_start_column, data_start_column + len(dates))):
            value = ws.cell(row=row_index, column=col).value
            if value is not None and isinstance(value, (int, float)):
                values[dates[date_idx]] = value
        
        # Create entry with hierarchy level
        entry = {
            "name": str(cell_value).strip(),
            "hierarchy_level": level,
            "row": row_index,
            "values": values
        }
        
        # Split name into Icelandic and English if available
        if " / " in entry["name"]:
            parts = entry["name"].split(" / ", 1)
            entry["is"] = parts[0]
            entry["en"] = parts[1]
        else:
            entry["is"] = entry["name"]
            entry["en"] = ""
            
        liabilities_data.append(entry)
        row_index += 1
    
    # Build hierarchical structure
    return build_hierarchy(excel_path, sheet_name, assets_data, liabilities_data, dates)

def build_hierarchy(excel_path, sheet_name, assets_data, liabilities_data, dates):
    """Build a hierarchical structure from the flat data"""
    result = {
        "metadata": {
            "file": os.path.basename(excel_path),
            "sheet": sheet_name,
            "dates": dates,
            "extracted_at": datetime.now().isoformat()
        },
        "assets": build_section_hierarchy(assets_data),
        "liabilities": build_section_hierarchy(liabilities_data)
    }
    return result

def build_section_hierarchy(section_data):
    """Build a hierarchical structure for a section (assets or liabilities)"""
    # Create a flat list of all items with their hierarchy info
    nodes = []
    
    for item in section_data:
        # Create node 
        node = {
            "id": f"row_{item['row']}",
            "name": item["name"],
            "is": item["is"],
            "en": item["en"],
            "level": item["hierarchy_level"],
            "values": item["values"],
            "children": []
        }
        nodes.append(node)
    
    # Build tree by iterating through the nodes
    root_nodes = []
    node_map = {}  # map of level arrays to track parents at each level
    
    for node in nodes:
        level = node["level"]
        
        if level == 1:
            # Top level node - add to root
            root_nodes.append(node)
            # Clear levels above this one
            node_map = {1: node}
        else:
            # Find the right parent - it's the latest node at the level above this one
            parent_level = level - 1
            if parent_level in node_map:
                parent = node_map[parent_level]
                parent["children"].append(node)
                # Set this as the latest node at this level
                node_map[level] = node
                # Clear any levels deeper than this one
                keys_to_remove = [k for k in node_map.keys() if k > level]
                for k in keys_to_remove:
                    node_map.pop(k, None)
    
    return root_nodes

def map_test_schema_to_hierarchy(excel_path="backend/cache/INN_ReikningarBankakerfis_012025.xlsx", output_path="backend/parsed_data/test_vs.json"):
    """Main function to map test.json schema to hierarchical JSON"""
    schema = load_schema()
    banking_config = schema["INN"]["data"]["reikningar_bankakerfis"]["config"]["sheets"][0]
    
    # Extract hierarchical data
    hierarchy_data = extract_hierarchy_data(excel_path)
    
    # Create a simplified readable structure
    readable_structure = {
        "metadata": {
            "file": os.path.basename(excel_path),
            "sheet": banking_config["sheet"],
            "schema_file": "test.json",
            "description": "Mapping of banking system accounts hierarchy with paths",
            "extracted_at": datetime.now().isoformat()
        },
        "file_info": {
            "is": schema["INN"]["is"],
            "en": schema["INN"]["en"],
            "data_type": schema["INN"]["data"]["reikningar_bankakerfis"]["is"],
            "data_type_en": schema["INN"]["data"]["reikningar_bankakerfis"]["en"]
        },
        "assets": extract_readable_hierarchy(hierarchy_data["assets"]),
        "liabilities": extract_readable_hierarchy(hierarchy_data["liabilities"])
    }
    
    # Save to file
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(readable_structure, f, indent=2, ensure_ascii=False)
    
    print(f"Saved hierarchical mapping to {output_path}")
    return readable_structure

def extract_readable_hierarchy(hierarchy, prefix="", path=None):
    """Extract a readable hierarchy with paths"""
    if path is None:
        path = []
    
    result = {}
    
    for i, item in enumerate(hierarchy):
        # Create a path identifier
        current_path = path + [i]
        path_str = ".".join(map(str, current_path))
        
        # Add to result
        result[path_str] = {
            "id": item.get("id", ""),
            "name": item.get("name", ""),
            "is": item.get("is", ""),
            "en": item.get("en", ""),
            "path": current_path,
            "level": len(current_path),
            "hierarchy_level": item.get("level", 0),
            "index": i,
            "full_path": prefix + (f"{item.get('name', '')}" if not prefix else f" > {item.get('name', '')}")
        }
        
        # Process children if any
        if "children" in item and item["children"]:
            child_prefix = prefix + (f"{item.get('name', '')}" if not prefix else f" > {item.get('name', '')}")
            child_results = extract_readable_hierarchy(item["children"], child_prefix, current_path)
            result.update(child_results)
    
    return result

if __name__ == "__main__":
    excel_path = "backend/cache/INN_ReikningarBankakerfis_012025.xlsx"
    output_path = "backend/parsed_data/test_vs.json"
    
    try:
        schema = load_schema()
        banking_config = schema["INN"]["data"]["reikningar_bankakerfis"]["config"]["sheets"][0]
        
        print(f"Mapping test.json schema to hierarchical JSON...")
        result = map_test_schema_to_hierarchy(excel_path, output_path)
        
        # Print summary
        print(f"Successfully created hierarchical mapping")
        print(f"Assets: {len(result['assets'])} items")
        print(f"Liabilities: {len(result['liabilities'])} items")
        
    except Exception as e:
        import traceback
        print(f"Error: {str(e)}")
        traceback.print_exc() 