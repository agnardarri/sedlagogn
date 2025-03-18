import json
import os
import pandas as pd
import openpyxl
from datetime import datetime
import numpy as np

def load_schema():
    """Load the schema from test.json"""
    schema_path = "backend/config/schemas/test.json"
    with open(schema_path, 'r', encoding='utf-8') as f:
        schema = json.load(f)
    return schema

def parse_excel_to_dataframe(excel_path="backend/cache/INN_ReikningarBankakerfis_012025.xlsx"):
    """Parse Excel file using test.json schema and convert to dataframe"""
    schema = load_schema()
    banking_config = schema["INN"]["data"]["reikningar_bankakerfis"]["config"]["sheets"][0]
    
    # Open Excel file
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_name = banking_config["sheet"]
    ws = wb[sheet_name]
    
    # Extract configuration parameters
    date_row = banking_config["date_row"]
    data_start_column = banking_config["data_start_column"]
    data_name_column = banking_config["data_name_column"]
    assets_row = banking_config["assets_row"]
    liabilities_row = banking_config["liabilities_row"]
    assets_hierarchy = banking_config["assets_hierarchy"]
    liabilities_hierarchy = banking_config["liabilities_hierarchy"]
    
    # Extract dates from header row
    dates = []
    for col in range(data_start_column, ws.max_column + 1):
        date_value = ws.cell(row=date_row, column=col).value
        if date_value:
            if isinstance(date_value, datetime):
                dates.append(date_value)
            else:
                try:
                    dt = pd.to_datetime(date_value)
                    dates.append(dt)
                except:
                    # If it's not a date, just skip it
                    pass
    
    # Store all data rows
    all_data = []
    
    # Process assets
    parent_stack = {}  # Track parents at each level
    current_parents = {}  # Current parent at each level
    
    row_index = assets_row
    for level in assets_hierarchy:
        # Get cell value (name)
        name = ws.cell(row=row_index, column=data_name_column).value
        
        if name is None:
            # Skip empty rows
            row_index += 1
            continue
        
        # Split name into Icelandic and English
        name_is = name
        name_en = ""
        if isinstance(name, str) and " / " in name:
            parts = name.split(" / ", 1)
            name_is = parts[0]
            name_en = parts[1]
        
        # Store this item as the current parent for this level
        current_parents[level] = name
        
        # Clear parents at deeper levels
        levels_to_remove = [k for k in current_parents.keys() if k > level]
        for k in levels_to_remove:
            if k in current_parents:
                del current_parents[k]
        
        # Get the direct parent (level - 1)
        parent = current_parents.get(level - 1, None) if level > 1 else None
        
        # Split parent into Icelandic and English
        parent_is = parent
        parent_en = ""
        if isinstance(parent, str) and " / " in parent:
            parts = parent.split(" / ", 1)
            parent_is = parts[0]
            parent_en = parts[1]
        
        # Extract values for each date
        for date_idx, col in enumerate(range(data_start_column, data_start_column + len(dates))):
            value = ws.cell(row=row_index, column=col).value
            
            if value is not None and isinstance(value, (int, float)):
                # Create a row in the dataframe format
                row = {
                    'date': dates[date_idx],
                    'name': name,
                    'name_is': name_is,
                    'name_en': name_en,
                    'type': 'asset',
                    'parent': parent,
                    'parent_is': parent_is,
                    'parent_en': parent_en,
                    'hierarchy_level': level,
                    'value': value
                }
                all_data.append(row)
        
        row_index += 1
    
    # Process liabilities (same logic as assets)
    current_parents = {}  # Reset current parents
    
    row_index = liabilities_row
    for level in liabilities_hierarchy:
        # Get cell value (name)
        name = ws.cell(row=row_index, column=data_name_column).value
        
        if name is None:
            # Skip empty rows
            row_index += 1
            continue
        
        # Split name into Icelandic and English
        name_is = name
        name_en = ""
        if isinstance(name, str) and " / " in name:
            parts = name.split(" / ", 1)
            name_is = parts[0]
            name_en = parts[1]
        
        # Store this item as the current parent for this level
        current_parents[level] = name
        
        # Clear parents at deeper levels
        levels_to_remove = [k for k in current_parents.keys() if k > level]
        for k in levels_to_remove:
            if k in current_parents:
                del current_parents[k]
        
        # Get the direct parent (level - 1)
        parent = current_parents.get(level - 1, None) if level > 1 else None
        
        # Split parent into Icelandic and English
        parent_is = parent
        parent_en = ""
        if isinstance(parent, str) and " / " in parent:
            parts = parent.split(" / ", 1)
            parent_is = parts[0]
            parent_en = parts[1]
        
        # Extract values for each date
        for date_idx, col in enumerate(range(data_start_column, data_start_column + len(dates))):
            value = ws.cell(row=row_index, column=col).value
            
            if value is not None and isinstance(value, (int, float)):
                # Create a row in the dataframe format
                row = {
                    'date': dates[date_idx],
                    'name': name,
                    'name_is': name_is,
                    'name_en': name_en,
                    'type': 'liability',
                    'parent': parent,
                    'parent_is': parent_is,
                    'parent_en': parent_en,
                    'hierarchy_level': level,
                    'value': value
                }
                all_data.append(row)
        
        row_index += 1
    
    # Create DataFrame
    df = pd.DataFrame(all_data)
    
    # Sort by date, type, and name
    df = df.sort_values(['date', 'type', 'name'])
    
    return df

def save_dataframe(df, csv_path="backend/parsed_data/banking_flat_data.csv", excel_path="backend/parsed_data/banking_flat_data.xlsx"):
    """Save the dataframe to CSV and Excel"""
    # Create the directory if it doesn't exist
    os.makedirs(os.path.dirname(csv_path), exist_ok=True)
    
    # Save as CSV
    df.to_csv(csv_path, index=False)
    print(f"Saved CSV to {csv_path}")
    
    # Save as Excel
    df.to_excel(excel_path, index=False)
    print(f"Saved Excel to {excel_path}")

def main():
    try:
        print("Parsing Excel file according to test.json schema...")
        df = parse_excel_to_dataframe()
        
        # Print DataFrame info
        print(f"Created DataFrame with {len(df)} rows")
        print(f"Date range: {df['date'].min()} to {df['date'].max()}")
        print(f"Asset rows: {len(df[df['type'] == 'asset'])}")
        print(f"Liability rows: {len(df[df['type'] == 'liability'])}")
        print(f"Hierarchy levels: {sorted(df['hierarchy_level'].unique())}")
        
        # Show sample
        print("\nSample data:")
        print(df[['date', 'name_is', 'name_en', 'type', 'parent_is', 'hierarchy_level', 'value']].head())
        
        # Save to files
        save_dataframe(df)
        
    except Exception as e:
        import traceback
        print(f"Error: {str(e)}")
        traceback.print_exc()

if __name__ == "__main__":
    main() 